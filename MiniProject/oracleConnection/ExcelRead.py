# -*- coding:utf8 -*-
import cx_Oracle
from openpyxl import *

# FARMID = 21

path = "K://2020년 3월 연구실 자료//스마트 양식장  어류 데이터 관련 문서//환경데이터 (1)//진도 보전//진도 보전.xls.xlsx"

load_wb = load_workbook(path, data_only=True)
load_ws = load_wb["Sheet1"]

max_row = load_ws.max_row

connection = cx_Oracle.connect("etc_user03","dclab","114.70.93.73",encoding="utf-8")
cur = connection.cursor()

for x in range(3,10):
    # sensor 1  = 3,4
    # sensor 2 =  5,6
    # sensor 3 =  7,8
    sensor1_do = load_ws.cell(x,7).value
    sensor1_wt = load_ws.cell(x,8).value

    sensor1_psurec = 0
    sensor1_phrec = 0
    sensor1_nh4rec = 0
    sensor1_no2rec = 0

    # Data Feature
    # TANKID
    # FARMID
    # FISHID
    # STATE, YRCODE, SENSORDATE
    # Sensor_do
    # Sensor_wt
    # Sensor_psurec
    # Sensor_phrec
    # Sensor_no4rec
    # Sensor_no2rec

    sql_query = "INSERT INTO rec VALUES (recseq.NEXTVAL," \
                + "'"    + "SENSOR3"  + "'" + "," \
                + "'"    + "22"    + "'" + "," \
                + "'"    + "1"     + "'" +"," \
                +"NULL," + "null," + "SYSDATE," \
                + "'" + str(sensor1_do) + "'" + ","\
                + "'" + str(sensor1_wt) + "'" + "," \
                + "'" + str(sensor1_psurec) + "'" + "," \
                + "'" + str(sensor1_phrec)  + "'" + "," \
                + "'" + str(sensor1_nh4rec) + "'" + "," \
                + "'" + str(sensor1_no2rec) + "'" + "," \
                +"SYSDATE, 'sysadmin' ,SYSDATE, 'sysadmin')"

    print(sql_query)
    cur.execute(sql_query)

connection.commit()
cur.close()
